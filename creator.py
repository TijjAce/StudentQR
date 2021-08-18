import xlsxwriter
from os import makedirs
import QrMake
from openpyxl import load_workbook
book=load_workbook('élèves.xlsx')
sheet=book.active

column = sheet.columns

firstColumn=[cell.value for cell in next(column)]
print(firstColumn)
secondColumn=[cell.value for cell in next(column)]
print(secondColumn)
thirdColumn=[cell.value for cell in next(column)]
print(thirdColumn)








x=0
for f, b ,c in zip(firstColumn,secondColumn,thirdColumn):
    x=x+1
    path=(f'file:///sdcard/Documents/dossierfichier/{c}/')
    newFile=str(x)
    stringc=str(c)
    creator=(str(f+b+stringc+newFile+'.xlsx'))

    try:
        makedirs(f'./dossierfichier/{c}', exist_ok=True)
        
    except OSError():
        pass
        


    print(creator)
    QrMake.associationducode(creator,path)
    workbook=xlsxwriter.Workbook(f'./dossierfichier/{c}/{f+b} {stringc}.xlsx')
    worksheet=workbook.add_worksheet()
    worksheet.write('A1', f)
    worksheet.write('B1', b)
    worksheet.write('C1', c)
    workbook.close()





